import json
from openpyxl import load_workbook

WB_PATH = "Agreements Log - FY26-27 Template (Update with Cluster Name) Interim for all Clusters.xlsx"

GROUP_KEYWORDS = {
    "Contact": ["name", "contact", "phone", "email", "address"],
    "Dates": ["date", "start", "end", "due", "received"],
    "Agreement": [
        "agreement",
        "agreement no",
        "agreement#",
        "agreement id",
        "contract",
        "id",
    ],
    "Location": ["cluster", "region", "location", "site"],
    "Financial": ["amount", "value", "budget", "cost", "fund"],
    "LOA": ["loa", "letter of agreement", "loa tracking"],
    "Notes": ["note", "comments", "remarks", "description"],
}


def choose_group(header):
    h = header.lower()
    for group, keywords in GROUP_KEYWORDS.items():
        for kw in keywords:
            if kw in h:
                return group
    return "Other"


def propose_mandatory(headers):
    mandatory = []
    candidates = ["name", "date", "cluster", "agreement", "id"]
    for h in headers:
        lh = h.lower()
        for c in candidates:
            if c in lh and h not in mandatory:
                mandatory.append(h)
    return mandatory


def main():
    wb = load_workbook(WB_PATH, read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]

    # Auto-detect header row: look in the first 10 rows for a row with several non-empty cells
    header_row_idx = 1
    for r in range(1, 11):
        row = [cell.value for cell in sheet[r]]
        non_empty = [c for c in row if c is not None and str(c).strip() != ""]
        if len(non_empty) >= 3 and not any(
            "agreements log" in str(c).lower() for c in non_empty
        ):
            header_row_idx = r
            break

    headers = []
    for cell in sheet[header_row_idx]:
        h = str(cell.value).strip() if cell.value is not None else ""
        if h == "__PowerAppsId__":
            continue
        headers.append(h)

    sample = {}
    first_data_row = None
    for row in sheet.iter_rows(
        min_row=header_row_idx + 1, max_row=header_row_idx + 1, values_only=True
    ):
        first_data_row = row
        break
    if first_data_row:
        for h, v in zip(headers, first_data_row):
            if h == "__PowerAppsId__":
                continue
            sample[h] = v

    groups = {}
    for h in headers:
        g = choose_group(h)
        groups.setdefault(g, []).append(h)

    out = {
        "sheet": sheet.title,
        "headers": headers,
        "sample_row": sample,
        "groups": groups,
        "proposed_mandatory": propose_mandatory(headers),
    }

    with open("field_groups.json", "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)

    print("Wrote field_groups.json")


if __name__ == "__main__":
    main()
