from openpyxl import load_workbook
import json

WB = "Agreements Log - FY26-27 Template (Update with Cluster Name) Interim for all Clusters.xlsx"
OUT = "field_lookups.json"


def main():
    wb = load_workbook(WB, data_only=True)
    # try to find the 'do not modify' sheet which holds lookup columns
    sheet = None
    for name in wb.sheetnames:
        if (
            "do not modify" in name.lower()
            or name.lower().replace(" ", "") == "donotmodify"
        ):
            sheet = wb[name]
            break
    if sheet is None:
        # fallback to first sheet
        sheet = wb[wb.sheetnames[0]]

    lookups = {}

    # dataValidations held on workbook. Try to read from sheet.data_validations
    try:
        dvs = list(sheet.data_validations.dataValidation)
    except Exception:
        dvs = []

    for dv in dvs:
        # dv.cells is a range like 'C2:C100' or similar
        try:
            sqref = dv.sqref
        except Exception:
            sqref = dv.cells
        # if dv.formula1 is a list like '"A,B,C"' or a reference to a range
        f = dv.formula1
        vals = None
        if f is None:
            continue
        if f.startswith('"') and f.endswith('"'):
            vals = [s.strip() for s in f.strip('"').split(",")]
        else:
            # assume it's a range like 'Sheet2'!$A$1:$A$10 or a named range
            # strip sheet prefix if present
            rng = f.replace("$", "")
            if "!" in rng:
                _, rr = rng.split("!")
            else:
                rr = rng
            rr = rr.strip('"')
            try:
                cells = wb.active[rr]
            except Exception:
                # try evaluate as named range
                cells = None
            if cells:
                vals = [c.value for c in cells]

        # assign lookup to the header in the same column if possible
        if vals:
            # find the column of the first sqref
            col = None
            if isinstance(sqref, str):
                # take first cell
                first = sqref.split(":")[0]
                # remove row number
                col = "".join([ch for ch in first if ch.isalpha()])
            if col:
                # find header value in that column in first 10 rows
                for r in range(1, 11):
                    cell = sheet[f"{col}{r}"]
                    if cell.value and str(cell.value).strip() != "":
                        header = str(cell.value).strip()
                        lookups[header] = [v for v in vals if v is not None]
                        break

    # Also check for named ranges that might hold lookups
    # iterate defined names properly
    for name in wb.defined_names:
        try:
            dests = list(wb.defined_names[name].destinations)
        except Exception:
            continue
        for sheetname, coord in dests:
            try:
                sh = wb[sheetname]
                cells = sh[coord]
                values = [c.value for c in cells]
                lookups[name] = [v for v in values if v is not None]
            except Exception:
                continue

    # If no direct data validations found, try heuristic: first, check the special lookup sheet
    if not lookups:
        # if there's a dedicated lookup sheet, use its first row as headers and columns as lists
        # find a sheet named like 'do not modify' (already assigned to 'sheet' above)
        # use that sheet as the lookup source only if its name indicates it's the lookup sheet
        lookup_sheet = None
        for name in wb.sheetnames:
            if (
                "do not modify" in name.lower()
                or name.lower().replace(" ", "") == "donotmodify"
            ):
                lookup_sheet = wb[name]
                break

        if lookup_sheet is not None:
            # first non-empty row as header row
            header_row = 1
            for r in range(1, min(lookup_sheet.max_row, 10) + 1):
                rowvals = [
                    lookup_sheet.cell(row=r, column=c).value
                    for c in range(1, lookup_sheet.max_column + 1)
                ]
                nonempty = [
                    v for v in rowvals if v is not None and str(v).strip() != ""
                ]
                if len(nonempty) >= 1:
                    header_row = r
                    break
            for c in range(1, lookup_sheet.max_column + 1):
                h = lookup_sheet.cell(row=header_row, column=c).value
                if not h:
                    continue
                # collect values below header
                vals = []
                for r in range(
                    header_row + 1, min(lookup_sheet.max_row, header_row + 1000) + 1
                ):
                    v = lookup_sheet.cell(row=r, column=c).value
                    if v is not None and str(v).strip() != "":
                        vals.append(str(v).strip())
                uniq = list(dict.fromkeys(vals))
                if uniq:
                    lookups[str(h).strip()] = uniq
        else:
            # fallback heuristic scanning other sheets for candidate lists
            candidates = {}
            for sname in wb.sheetnames:
                sh = wb[sname]
                # detect header row for this sheet (first row with >=2 non-empty cells)
                header_row = 1
                for r in range(1, min(sh.max_row, 20) + 1):
                    rowvals = [
                        sh.cell(row=r, column=c).value
                        for c in range(1, sh.max_column + 1)
                    ]
                    nonempty = [
                        v for v in rowvals if v is not None and str(v).strip() != ""
                    ]
                    if len(nonempty) >= 2:
                        header_row = r
                        break

                for col_idx in range(1, sh.max_column + 1):
                    vals = []
                    for r in range(
                        header_row + 1, min(sh.max_row, header_row + 200) + 1
                    ):
                        v = sh.cell(row=r, column=col_idx).value
                        if v is not None and str(v).strip() != "":
                            vals.append(str(v).strip())
                    # filter out values that look like headers or sheet titles
                    filtered = [
                        v
                        for v in vals
                        if v.upper()
                        not in (sh.title.upper(), "TOTAL", "INSERT ROWS ABOVE")
                    ]
                    uniq = list(dict.fromkeys(filtered))
                    if 1 < len(uniq) <= 200:
                        key = f"{sname}!{col_idx}"
                        candidates[key] = uniq

            # try to map candidates to headers in main sheet
            main = sheet
            header_row = 1
            for r in range(1, 11):
                rowvals = [
                    main.cell(row=r, column=c).value
                    for c in range(1, main.max_column + 1)
                ]
                nonempty = [
                    v for v in rowvals if v is not None and str(v).strip() != ""
                ]
                if len(nonempty) >= 2:
                    header_row = r
                    break
            for c in range(1, main.max_column + 1):
                h = main.cell(row=header_row, column=c).value
                if not h:
                    continue
                h = str(h).strip()
                # sample values below header
                sample_vals = []
                for r in range(header_row + 1, min(main.max_row, header_row + 200) + 1):
                    v = main.cell(row=r, column=c).value
                    if v is not None and str(v).strip() != "":
                        sample_vals.append(str(v).strip())
                if not sample_vals:
                    continue
                # compare against candidates
                for key, cand in candidates.items():
                    cand_set = set(cand)
                    match_count = sum(1 for v in sample_vals if v in cand_set)
                    if match_count / max(1, len(sample_vals)) >= 0.6:
                        lookups[h] = cand
                        break

    # remove PowerAppsId if found
    if "__PowerAppsId__" in lookups:
        del lookups["__PowerAppsId__"]

    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(lookups, f, indent=2, ensure_ascii=False)
    print("Wrote", OUT)


if __name__ == "__main__":
    main()
